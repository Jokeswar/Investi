import os
import json
import openpyxl
import variabiles as var

from datetime import datetime
from multiprocessing import Pool
from urllib.request import Request, urlopen
from bs4 import BeautifulSoup


infoJson = infoJson = json.load(open(os.path.join("data", "loan_originators_req.json")))


class Loan:
    link = ""
    originator = ""
    daysDue = 0.0
    interestRate = 0.0
    amount = 0.0
    amountAvailable = 0.0

    def __init__(self, link, originator, daysDue, interestRate, amount, amav):
        self.link = link
        self.originator = originator
        self.daysDue = daysDue
        self.interestRate = interestRate
        self.amount = amount
        self.amountAvailable = amav

    def __str__(self):
        return "{} | Interest Rate: {}% | Days Due: {} | Originator: {} | Ammount: {} | Available: {}\n".format(self.link, self.interestRate, self.daysDue, self.originator, self.amount, self.amountAvailable)


def GetInfoTable(link):
    result = []
    req = Request(link, headers={'User-Agent': 'Mozilla/5.0'})
    html = urlopen(req).read()
    soup = BeautifulSoup(html, 'html.parser')

    infoTable = soup.findAll("table", {"class": "value-table"})[-1].findChildren()[0]
    for tr in infoTable:
        result.append(tr.findChildren()[1].text)

    return result


# eval function for loans
def Pass(val, req, oper):
    if any(char.isdigit() for char in val) is False:
        val = "".join(val)

    if any(char.isdigit() for char in req) is False:
        req = "".join(req)

    return eval(oper.format(val, req))


def validate(loan):
    global infoJson
    infoTable = GetInfoTable(loan.link)

    operations = infoJson[loan.originator]["additionalInformationOperation"].split(",")
    infoRequirements = infoJson[loan.originator]["additionalInformationRequierement"].split(",")
    infoRows = infoJson[loan.originator]["additionalInformationRow"].split(",")

    valid = True
    info = []
    testValuesLen = len(operations)

    try:
        for i in range(0, testValuesLen):
            information = infoTable[int(infoRows[i])]  # Data from loan page
            info.append(information)

        for i in range(0, testValuesLen):
            if Pass(info[i], infoRequirements[i], operations[i]) is False:
                valid = False
                break

        if valid:
            return loan
    except:
        print("Failed to validate loan: " + loan.link)

    return None


def analyze():
    infoJson = json.load(open(var.GLOBALS["JSONREQ"]))

    countryMaxInv = {}
    loans = []

    workbook = openpyxl.load_workbook(var.GLOBALS["LOANDATA"])
    worksheet = workbook.active

    # Setting maximum loan amount for every country
    for loanOriginator in infoJson["LoanOriginators"]:
        wage = float(infoJson[loanOriginator]["wage"])
        DTI = float(infoJson[loanOriginator]["DTI"]) # Debt to income
        countryMaxInv[infoJson[loanOriginator]["mintosCountryName"]] = wage * DTI

    for rowN in range(2, worksheet.max_row):
        # Get necessary data
        country = str(worksheet.cell(row=rowN, column=1).value)
        loanId = str(worksheet.cell(row=rowN, column=2).value)
        dueDateValue = str(worksheet.cell(row=rowN, column=4).value)
        loanOriginator = str(worksheet.cell(row=rowN, column=7).value)
        loanAmmount = float(worksheet.cell(row=rowN, column=9).value)
        interestRate = float(worksheet.cell(row=rowN, column=12).value)
        amountAvailable = float(worksheet.cell(row=rowN, column=16).value)

        # Get date
        day = int(dueDateValue.split(".")[0])
        month = int(dueDateValue.split(".")[1])
        year = int(dueDateValue.split(".")[2])
        dueDate = datetime(year, month, day)

        daysDue = (dueDate - datetime.now()).days + 1

        if loanOriginator in infoJson["LoanOriginators"]:
            if loanAmmount < countryMaxInv[country]:
                link = "https://www.mintos.com/en/" + str(loanId)
                newLoan = Loan(link, loanOriginator, daysDue, interestRate, loanAmmount, amountAvailable)
                loans.append(newLoan)

    loans.sort(key=lambda loan: (loan.interestRate, loan.daysDue), reverse=True)

    # output = open(OUTPUTFILE, "w")

    with Pool(processes = os.cpu_count()) as pool:
        validatedLoans = pool.map(validate, loans)

    # for loan in validatedLoans:
    #     if loan != None:
    #         output.write(str(loan))

    # output.close()

    return validatedLoans
